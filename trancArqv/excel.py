from openpyxl import *


class LerExcel:

   def __init__(self, arquivo, planilha):
       self.wb = load_workbook(arquivo)
       self.ws = self.wb[planilha]

   def cell_string(self, linha, coluna):
       dado = self.ws.cell(column=coluna, row=linha).value
       return str(dado)

   def cell_int(self, linha, coluna):
       dado = self.ws.cell(column=coluna, row=linha).value
       try:
           int(dado)
           return int(dado)
       except:
           return -999999999

   def cell_float(self, linha, coluna):
       dado = self.ws.cell(column=coluna, row=linha).value
       try:
           float(dado)
           return float(dado)
       except:
           return -999999999

   def get_max_linha(self):
       return self.ws.max_row

   def get_max_coluna(self):
       return self.ws.max_column
